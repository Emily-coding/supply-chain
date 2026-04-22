"""
Compute observed (empirical) energy-price pass-through ρⱼ from ONS PPI,
and ONS CPI response, for the 2021 Q1 → 2022 Q4 shock window.

Steps:
  1. Open ppi.xlsx once, extract all PPI INDEX OUTPUT DOMESTIC series for
     monthly rows 1122–1169 (2021-01 .. 2024-12).
  2. Parse the CPA product code out of each title, aggregate series by
     IOAT industry code, and compute mean PPI index.
  3. Extract the energy PPI series (D351 elec, D352 gas, C19 refined)
     and compute the realised energy shock between 2021-Jan and 2022-peak.
  4. Combine with Leontief energy shares (from irf.json) to compute the
     theoretical full pass-through and back out empirical ρⱼ.
  5. Open mm23.xlsx, extract CPI divisions, compute actual consumer-price
     moves over the same window.
  6. Write data/observed.json for the HTML page.
"""
import json, re, time
from pathlib import Path
import numpy as np
import openpyxl

ROOT = Path(__file__).parent
PPI  = ROOT / "data" / "ppi.xlsx"
CPI  = ROOT / "data" / "mm23.xlsx"
IRF  = ROOT / "data" / "irf.json"
OUT  = ROOT / "data" / "observed.json"

# Monthly data layout:
# 2012-01 = r1014, so YYYY-MM = r1014 + (YYYY-2012)*12 + (MM-1)
def month_row(y, m): return 1014 + (y-2012)*12 + (m-1)

SHOCK_START = (2021, 1)
SHOCK_END   = (2022, 10)   # ~UK PPI peak; gas peaked Aug 2022 but PPI lags
WINDOW_END  = (2024, 12)

# PPI series -> IOAT industry mapping.
# Multiple CPA series may map to one IOAT industry; they will be averaged.
# Keys on left = CPA code prefix as it appears in PPI titles.
# Values on right = IOAT code from irf.json.
CPA_TO_IOAT = {
    # Mining / fuels
    "B0510":"B05","B0520":"B05","B0610":"B06 & B07","B0620":"B06 & B07",
    "B0710":"B06 & B07","B0721":"B06 & B07","B0729":"B06 & B07",
    "B0812":"B08","B0892":"B08","B0893":"B08","B0899":"B08",
    "B0910":"B09","B0990":"B09",
    # Food
    "C1011":"C101","C1012":"C101","C1013":"C101",
    "C1020":"C102_3",
    "C1031":"C102_3","C1032":"C102_3","C1039":"C102_3",
    "C1041":"C104","C1042":"C104",
    "C1051":"C105","C1052":"C105",
    "C1061":"C106","C1062":"C106",
    "C1071":"C107","C1072":"C107","C1073":"C107",
    "C1081":"C108","C1082":"C108","C1083":"C108","C1084":"C108","C1085":"C108","C1086":"C108","C1089":"C108",
    "C1091":"C109","C1092":"C109",
    # Beverages + tobacco
    "C1101":"C1101T1106 & C12","C1102":"C1101T1106 & C12","C1103":"C1101T1106 & C12",
    "C1104":"C1101T1106 & C12","C1105":"C1101T1106 & C12","C1106":"C1101T1106 & C12",
    "C1107":"C1107",
    "C1200":"C1101T1106 & C12",
    # Textiles
    "C131":"C13","C132":"C13","C133":"C13","C1391":"C13","C1392":"C13","C1393":"C13","C1394":"C13","C1395":"C13","C1396":"C13","C1399":"C13",
    "C1411":"C14","C1412":"C14","C1413":"C14","C1414":"C14","C1419":"C14","C142":"C14","C143":"C14",
    "C1511":"C15","C1512":"C15","C1520":"C15",
    # Wood, paper, printing
    "C1610":"C16","C1621":"C16","C1622":"C16","C1623":"C16","C1624":"C16","C1629":"C16",
    "C1711":"C17","C1712":"C17","C1721":"C17","C1722":"C17","C1723":"C17","C1724":"C17","C1729":"C17",
    "C1811":"C18","C1812":"C18","C1813":"C18","C1814":"C18","C1820":"C18",
    # Coke and refined petroleum
    "C1910":"C19","C1920":"C19",
    # Chemicals — split across C203/204/205/20A/20B/20C in IxI
    "C2011":"C20A","C2012":"C20C","C2013":"C20A","C2014":"C20B","C2015":"C20A",
    "C2016":"C20B","C2017":"C20B","C2020":"C20C",
    "C2030":"C203",
    "C2041":"C204","C2042":"C204",
    "C2051":"C205","C2052":"C205","C2053":"C205","C2059":"C205",
    "C2060":"C20B",
    # Pharmaceuticals
    "C2110":"C21","C2120":"C21",
    # Rubber & plastics
    "C2211":"C22","C2219":"C22","C2221":"C22","C2222":"C22","C2223":"C22","C2229":"C22",
    # Non-metallic minerals
    "C2311":"C23OTHER","C2312":"C23OTHER","C2313":"C23OTHER","C2314":"C23OTHER","C2319":"C23OTHER","C2320":"C23OTHER",
    "C2331":"C23OTHER","C2332":"C23OTHER",
    "C2341":"C23OTHER","C2342":"C23OTHER","C2343":"C23OTHER","C2344":"C23OTHER","C2349":"C23OTHER",
    "C2351":"C235_6","C2352":"C235_6","C2361":"C235_6","C2362":"C235_6","C2363":"C235_6","C2364":"C235_6",
    "C2365":"C235_6","C2369":"C235_6","C2370":"C23OTHER","C2391":"C23OTHER","C2399":"C23OTHER",
    # Metals
    "C2410":"C241T243","C2420":"C241T243","C2431":"C241T243","C2432":"C241T243","C2433":"C241T243","C2434":"C241T243",
    "C2441":"C244_5","C2442":"C244_5","C2443":"C244_5","C2444":"C244_5","C2445":"C244_5","C2446":"C244_5","C2451":"C244_5","C2452":"C244_5","C2453":"C244_5","C2454":"C244_5",
    # Fabricated metal
    "C2511":"C25","C2512":"C25","C2521":"C25","C2529":"C25","C2530":"C25","C2540":"C25","C2550":"C25",
    "C2561":"C25","C2562":"C25","C2571":"C25","C2572":"C25","C2573":"C25","C2591":"C25","C2592":"C25","C2593":"C25","C2594":"C25","C2599":"C25",
    # Computer, electronic, optical
    "C2611":"C26","C2612":"C26","C2620":"C26","C2630":"C26","C2640":"C26","C2651":"C26","C2652":"C26","C2660":"C26","C2670":"C26","C2680":"C26",
    # Electrical equipment
    "C2711":"C27","C2712":"C27","C2720":"C27","C2731":"C27","C2732":"C27","C2733":"C27","C2740":"C27","C2751":"C27","C2752":"C27","C2790":"C27",
    # Machinery
    "C2811":"C28","C2812":"C28","C2813":"C28","C2814":"C28","C2815":"C28","C2821":"C28","C2822":"C28","C2823":"C28",
    "C2824":"C28","C2825":"C28","C2829":"C28","C2830":"C28","C2841":"C28","C2849":"C28","C2891":"C28","C2892":"C28","C2893":"C28","C2894":"C28","C2895":"C28","C2896":"C28","C2899":"C28",
    # Motor vehicles and transport
    "C2910":"C29","C2920":"C29","C2931":"C29","C2932":"C29",
    "C3011":"C301","C3012":"C301",
    "C3030":"C303",
    "C3020":"C30OTHER","C3091":"C30OTHER","C3092":"C30OTHER","C3099":"C30OTHER",
    # Furniture + other manufacturing
    "C3101":"C31","C3102":"C31","C3103":"C31","C3109":"C31",
    "C3211":"C32","C3212":"C32","C3213":"C32","C3220":"C32","C3230":"C32","C3240":"C32","C3250":"C32","C3291":"C32","C3299":"C32",
    # Repair & install
    "C3315":"C3315","C3316":"C3316",
    "C3311":"C33OTHER","C3312":"C33OTHER","C3313":"C33OTHER","C3314":"C33OTHER","C3317":"C33OTHER","C3319":"C33OTHER","C3320":"C33OTHER",
    # Electricity and gas (D)
    "D3511":"D351","D3512":"D351","D3513":"D351","D3514":"D351",
    "D3521":"D352_3","D3522":"D352_3","D3523":"D352_3",
}

# -----------------------------------------------------------------------------
def parse_cpa(title: str):
    """Extract the CPA/SIC code token from an ONS PPI title."""
    if not title or " - " not in title: return None
    tail = title.split(" - ", 1)[1]
    m = re.match(r'([A-Z0-9_]+(?:\.[0-9]+)?)\s', tail)
    if not m: return None
    return m.group(1).rstrip(".")


def load_ppi():
    """Extract all PPI OUTPUT DOMESTIC monthly values 2021-01 .. 2024-12."""
    print(f"loading PPI …", flush=True); t0 = time.time()
    wb = openpyxl.load_workbook(PPI, data_only=True)
    ws = wb["data"]
    print(f"  loaded in {time.time()-t0:.1f}s", flush=True)

    # Build the month → row index we need (48 months)
    months = []
    for y in range(2021, 2025):
        for m in range(1, 13):
            months.append((y, m, month_row(y, m)))

    # Scan columns: we want "PPI INDEX OUTPUT DOMESTIC" series, non-duplicate
    # (prefer "excluding Duty/Climate Change Levy" for cleaner pass-through signal).
    series = []
    for c in range(2, ws.max_column + 1):
        t = ws.cell(1, c).value
        if not t or not t.startswith("PPI INDEX OUTPUT DOMESTIC - "):
            continue
        # Skip "including CCL/Duty" duplicates
        if "including" in t.lower():
            continue
        cpa = parse_cpa(t)
        if cpa is None:
            continue
        cdid = ws.cell(2, c).value
        series.append({"col": c, "cpa": cpa, "cdid": cdid, "title": t})

    print(f"  {len(series)} OUTPUT DOMESTIC series (excl duplicates)", flush=True)

    # Read monthly values for each series
    for s in series:
        s["values"] = {}
        for y, m, r in months:
            v = ws.cell(r, s["col"]).value
            if isinstance(v, (int, float)):
                s["values"][f"{y}-{m:02d}"] = float(v)

    wb.close()
    return series


def load_cpi():
    """Extract monthly CPI ALL-ITEMS and main divisions, 2021-01 .. 2024-12."""
    print(f"loading CPI …", flush=True); t0 = time.time()
    wb = openpyxl.load_workbook(CPI, data_only=True)
    ws = wb["data"]
    print(f"  loaded in {time.time()-t0:.1f}s", flush=True)

    # CPI file layout: same as PPI (col A row 8+ time labels). But monthly rows
    # and row offsets may differ — scan col A around row 1100 to locate.
    # Find "2021 JAN"
    r_2021_01 = None
    for r in range(900, ws.max_row + 1):
        if ws.cell(r, 1).value == "2021 JAN":
            r_2021_01 = r
            break
    if r_2021_01 is None:
        raise SystemExit("cannot locate '2021 JAN' row in mm23.xlsx")

    months = []
    for i, (y, m) in enumerate([(y, m) for y in range(2021, 2025) for m in range(1, 13)]):
        months.append((y, m, r_2021_01 + i))

    # Target CPI titles. CPI level series (2015=100) rather than rate series.
    targets = {
        "all":         ["cpi index 00", "all items"],
        "food":        ["cpi index 01", "food"],
        "energy":      ["cpi index 04.5", "electricity, gas"],
        "housing":     ["cpi index 04", "housing, water"],
        "transport":   ["cpi index 07", "transport"],
        "restaurants": ["cpi index 11", "restaurants"],
    }
    # Scan columns: find first column whose lowered title starts with "cpi index"
    # and contains the target division key; prefer non-rate, non-percent series.
    found = {}
    for c in range(2, ws.max_column + 1):
        t = ws.cell(1, c).value
        if not isinstance(t, str): continue
        tl = t.lower()
        if "annual rate" in tl or "monthly rate" in tl or "percentage" in tl:
            continue
        for key, needles in targets.items():
            if key in found: continue
            if all(n in tl for n in needles):
                found[key] = {"col": c, "title": t, "cdid": ws.cell(2,c).value, "values":{}}
                break

    # Read values
    for key, s in found.items():
        for y, m, r in months:
            v = ws.cell(r, s["col"]).value
            if isinstance(v, (int, float)):
                s["values"][f"{y}-{m:02d}"] = float(v)

    print(f"  found {len(found)} CPI series", flush=True)
    wb.close()
    return found


# -----------------------------------------------------------------------------
def aggregate_by_ioat(series):
    """Average PPI series by IOAT industry code. Return {ioat: {month: avg}}."""
    # Division-level fallback map, used when a precise 4-digit CPA isn't listed.
    DIVISION_FALLBACK = {
        "A02":"A02", "A03":"A03", "A01":"A01",
        "B05":"B05", "B06":"B06 & B07", "B07":"B06 & B07", "B08":"B08", "B09":"B09",
        "C10":"C108", "C11":"C1101T1106 & C12", "C12":"C1101T1106 & C12",
        "C13":"C13","C14":"C14","C15":"C15","C16":"C16","C17":"C17","C18":"C18",
        "C19":"C19","C20":"C205","C21":"C21","C22":"C22",
        "C23":"C23OTHER",
        "C24":"C241T243",
        "C25":"C25","C26":"C26","C27":"C27","C28":"C28",
        "C29":"C29","C30":"C30OTHER","C31":"C31","C32":"C32","C33":"C33OTHER",
        "D35":"D351",
        "E36":"E36","E37":"E37","E38":"E38","E39":"E39",
    }
    buckets = {}
    unmapped_cpa = {}
    for s in series:
        cpa = s["cpa"]
        ioat = CPA_TO_IOAT.get(cpa)
        if ioat is None:
            for trunc_len in (5,4,3):
                ioat = CPA_TO_IOAT.get(cpa[:trunc_len])
                if ioat: break
        if ioat is None:
            # Division fallback: first 3 chars like "C20", "E36"
            ioat = DIVISION_FALLBACK.get(cpa[:3])
        if ioat is None:
            unmapped_cpa[cpa] = unmapped_cpa.get(cpa, 0) + 1
            continue
        buckets.setdefault(ioat, []).append(s)

    # For each bucket compute average monthly index (base = mean of 2021 H1 = 100)
    out = {}
    for ioat, slist in buckets.items():
        # Normalise each series to its own 2021-01 value, then average
        normed = []
        for s in slist:
            v0 = s["values"].get("2021-01")
            if not v0: continue
            normed.append({m: v / v0 * 100 for m, v in s["values"].items()})
        if not normed: continue
        # Compute month-wise mean (handling missing)
        all_months = set()
        for n in normed: all_months.update(n.keys())
        avg = {}
        for m in sorted(all_months):
            xs = [n[m] for n in normed if m in n]
            if xs: avg[m] = sum(xs) / len(xs)
        out[ioat] = {
            "index": avg,               # {month: PPI rebased to 2021-01 = 100}
            "n_series": len(slist),
            "cpas": [s["cpa"] for s in slist][:8],
        }

    if unmapped_cpa:
        print(f"  warning: {len(unmapped_cpa)} CPA codes unmapped, e.g. "
              f"{list(unmapped_cpa.items())[:8]}")
    return out


def compute_pass_through(ppi_by_ioat, irf_data):
    """For each industry, compute observed ρ and compare to the Leontief prediction."""
    # Energy shock from PPI
    if "C19" not in ppi_by_ioat or "D351" not in ppi_by_ioat or "D352_3" not in ppi_by_ioat:
        raise SystemExit("missing energy PPI series after aggregation")

    def idx(ioat, ym):
        return ppi_by_ioat[ioat]["index"].get(f"{ym[0]}-{ym[1]:02d}")

    # Measure peak shock for each energy industry over the window (2021-Jan → peak month in 2021-24)
    energy = {}
    for ec in ("C19","D351","D352_3"):
        series = ppi_by_ioat[ec]["index"]
        v0 = series.get("2021-01", 100.0)
        peak_m = max(series, key=lambda m: series[m])
        peak_v = series[peak_m]
        avg_22 = np.mean([series.get(f"2022-{m:02d}", np.nan) for m in range(1,13)])
        energy[ec] = {
            "p0": v0, "peak_month": peak_m, "peak_v": peak_v,
            "peak_shock": peak_v/v0 - 1,
            "avg_2022_shock": avg_22/v0 - 1 if v0 else None,
        }

    # Build energy-shock vector (average 2022 shock per series)
    # Use peak shock as the headline, but compute ρ at peak month per industry
    rows_out = []
    industries = {i["code"]: i for i in irf_data["industries"]}

    for code, rec in industries.items():
        if code not in ppi_by_ioat:
            continue
        ppi = ppi_by_ioat[code]["index"]
        v0 = ppi.get("2021-01", 100.0)
        # Observed industry PPI moves at key horizons
        obs_q1  = (ppi.get("2021-04", None)/v0 - 1) if ppi.get("2021-04") else None
        obs_q4  = (ppi.get("2022-01", None)/v0 - 1) if ppi.get("2022-01") else None
        obs_q8  = (ppi.get("2023-01", None)/v0 - 1) if ppi.get("2023-01") else None
        obs_q12 = (ppi.get("2024-01", None)/v0 - 1) if ppi.get("2024-01") else None
        # Peak observed move in the window
        peak_m = max(ppi, key=lambda m: ppi[m])
        obs_peak = ppi[peak_m]/v0 - 1

        # Theoretical full-pass-through ΔPPI from the industry's A coefficients
        # × observed energy shocks (use avg 2022 as "sustained" shock).
        a_gas  = rec["direct_gas"]
        a_elec = rec["direct_elec"]
        a_oil  = rec["direct_oil"]
        s_gas  = energy["D352_3"]["avg_2022_shock"] or 0.0
        s_elec = energy["D351"]["avg_2022_shock"] or 0.0
        s_oil  = energy["C19"]["avg_2022_shock"] or 0.0

        # Direct-only theoretical
        theor_direct = a_gas*s_gas + a_elec*s_elec + a_oil*s_oil
        # Full-Leontief theoretical move under the observed 2022-avg shock.
        # unit_full_lr gives the long-run response to +10% on every energy
        # product; rescale linearly by the avg energy shock observed.
        avg_shock = (s_gas + s_elec + s_oil) / 3.0
        theor_full = (rec["unit_full_lr"] / 0.10) * avg_shock if avg_shock else 0.0

        # Empirical pass-through: observed ÷ Leontief-full theoretical.
        # rho = 1 means the PPI moved exactly as the Leontief model (with full
        # pass-through) predicts from the observed energy shock alone.
        # rho > 1 means the PPI rose more than energy alone explains — either
        # other cost shocks (wages, non-energy inputs) or excess pass-through
        # (LMM market-power result). rho < 1 means partial pass-through /
        # margin compression / output adjustment instead of price.
        rho_full   = (obs_q8 / theor_full)   if (obs_q8 is not None and theor_full)   else None
        rho_direct = (obs_q8 / theor_direct) if (obs_q8 is not None and theor_direct) else None
        rho_peak   = (obs_peak / theor_full) if (theor_full) else None

        rows_out.append({
            "code": code,
            "name": rec["name"],
            "n_ppi_series": ppi_by_ioat[code]["n_series"],
            "direct_share": rec["direct_share"],
            "theor_direct_22": theor_direct,
            "theor_full_22":   theor_full,
            "obs_q1":   obs_q1,
            "obs_q4":   obs_q4,
            "obs_q8":   obs_q8,
            "obs_q12":  obs_q12,
            "obs_peak": obs_peak,
            "peak_month": peak_m,
            "rho_direct_q8": rho_direct,
            "rho_full_q8":   rho_full,
            "rho_peak":      rho_peak,
            "rho_assumed":   rec["rho"],
        })

    return rows_out, energy


# -----------------------------------------------------------------------------
def main():
    irf = json.loads(IRF.read_text())

    ppi_series = load_ppi()
    ppi_by_ioat = aggregate_by_ioat(ppi_series)
    print(f"  {len(ppi_by_ioat)} IOAT industries matched to PPI series")

    cpi_series = load_cpi()

    rows, energy_shock = compute_pass_through(ppi_by_ioat, irf)

    # Assemble output
    result = {
        "window":    {"start":"2021-01", "end":"2024-12"},
        "shock_window": {"start":"2021-01", "peak_end": max(energy_shock[k]['peak_month'] for k in energy_shock)},
        "energy_ppi": {k: {kk:vv for kk,vv in v.items() if kk!='peak_month'} |
                           {"peak_month": v["peak_month"]} for k,v in energy_shock.items()},
        "industries": rows,
        "cpi": {k: {
                    "title":v["title"], "cdid":v["cdid"],
                    "index_norm":{m: val/v["values"].get("2021-01",val)*100
                                   for m,val in v["values"].items() if v["values"].get("2021-01")}
               } for k,v in cpi_series.items()},
        "ppi_index_by_ioat": {k: v["index"] for k, v in ppi_by_ioat.items()},
        "notes": (
            "Observed monthly PPI Output Domestic indices (ONS MM22, release "
            "2026-04-22) aggregated to the 101 IOAT industries of the 2023 "
            "analytical table, rebased to 2021-01 = 100. Empirical pass-through "
            "ρ_direct = obs_q8 / (direct energy share × avg-2022 energy shock)."
        ),
    }
    OUT.write_text(json.dumps(result, indent=1))
    print(f"wrote {OUT}  ({len(rows)} industries, {len(cpi_series)} CPI series)")


if __name__ == "__main__":
    main()
