"""
UK Supply-Chain Leontief energy exposures + price IRF.

Reads the 2023 ONS Input-Output Analytical Tables (industry-by-industry) from
`data/iot2023industry.xlsx` and, for each non-energy industry, emits two
distinct things:

1. Structural energy exposures per £1 of output, using a corrected
   energy-intensity vector at each domestic node:
       e_fuel[i] = A_dom[fuel, i] + A_imp[fuel, i]   (UK + imports)
   for fuel ∈ {gas, electricity, refined petroleum}, where A_imp comes
   from the ONS 'Imports use pxi' table. The Leontief inverse remains
   strictly domestic, L_dom = (I − A_dom)⁻¹. Reported per industry j:
   - Direct  = e_fuel[j]                         (UK + imp energy bought directly)
   - Total   = e_fuel · L_dom[:, j]              (embedded via UK supply chains)
   - Indirect = Total − Direct
   - Indirect oil split by last-hop attribution:
       oil_via_transport          = Σ_{T ∈ transport} e_oil[T] · L_dom[T, j]
       oil_indirect_non_transport = Indirect_oil − oil_via_transport
     where "transport" = every IOAT 'H' bucket (rail, land transport,
     water, air, warehousing, post/courier).
   This captures imported jet/bunker/diesel both directly (for industries
   that buy them) and via the transport channel (for industries that buy
   transport services). It does NOT capture foreign embodied energy in
   imported NON-energy intermediates (e.g. Chinese steel, German chemicals).
   The per-industry import_intensity field flags where that residual gap
   is most likely to matter.

2. A price-side impulse response: treat C19, D351, D352_3 as exogenous
   price-setters and propagate shocks through the rest of the economy.
   The energy → non-energy block of A uses the SAME corrected energy-
   intensity vector e_fuel[j] = A_dom[fuel,j] + A_imp[fuel,j] (so an
   industry that buys imported jet kerosene gets a bigger price response
   than the domestic-only IRF would predict). The non-energy → non-energy
   block stays domestic-A. Two scenarios — UNIT (+10% on every energy
   product, an elasticity probe) and STRESS (oil +100%, elec +50%,
   gas +50% — a forward-looking oil-led disturbance, comparable in
   magnitude to 2022 but with the lead-energy reversed). Two pass-through
   regimes — FULL
   (ρ=1) and REALISTIC (sector-specific ρ_j, calibrated to Ganapati/
   Shapiro/Walker 2020 and Lafrogne-Joussier/Martin/Méjean 2023). Dynamic
   path built by Neumann series; one expansion round = one quarter.

Output: data/irf.json, consumed by index.html and summary.html.
"""
import json
from pathlib import Path
import numpy as np
import openpyxl

ROOT = Path(__file__).parent
XLSX = ROOT / "data" / "iot2023industry.xlsx"
XLSX_PRODUCT = ROOT / "data" / "iot2023product.xlsx"
OUT  = ROOT / "data" / "irf.json"

ENERGY_CODES = ["C19", "D351", "D352_3"]
# All UK transport / logistics industries — every IOAT 'H' bucket. Oil
# bought by these industries is the various fuels used to move goods and
# people (diesel for road haulage, jet kerosene for air, marine bunker for
# water, electricity-displacing diesel for rail traction, last-mile diesel
# for post/courier). For a downstream industry j, the indirect oil that
# arrives via "any transport last-hop" is summed across this set.
TRANSPORT_CODES = ["H491_2", "H493T495", "H50", "H51", "H52", "H53"]
HORIZON_QUARTERS = 12

# Scenario shocks (decimal, not pp).
#   UNIT  is the elasticity probe — +10% on each energy product simultaneously.
#   STRESS is a forward-looking oil-led disturbance, sized like a Strait of
#         Hormuz / OPEC+ severe-cut episode: refined oil doubles, gas and
#         electricity rise 50%. Comparable in magnitude to 2022 but with the
#         lead-energy reversed. Tunable here; downstream code reads them via
#         the scen_defs dict in build().
SHOCK_UNIT   = {"C19": 0.10, "D351": 0.10, "D352_3": 0.10}
SHOCK_STRESS = {"C19": 1.00, "D351": 0.50, "D352_3": 0.50}

# -----------------------------------------------------------------------------
# Load GVA + total output per industry from the IOT sheet of the same workbook.
# Used to convert energy cost shares from a share-of-output denominator (the
# A-matrix native unit) to a share-of-GVA denominator (the ETII / DESNZ
# convention, and how most of the cost-pass-through literature reports
# energy intensity).
# Layout: IOT sheet, columns 3..N are industries (codes at r4), GVA at r118,
# total output at basic prices at r119.
# -----------------------------------------------------------------------------
def load_gva_and_output():
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    ws = wb["IOT"]
    out = {}
    for c in range(3, ws.max_column + 1):
        code = ws.cell(4, c).value
        if code is None:
            break
        code = str(code).strip()
        gva = ws.cell(118, c).value
        output = ws.cell(119, c).value
        out[code] = {
            "gva_gbpm":    float(gva)    if isinstance(gva,    (int, float)) else 0.0,
            "output_gbpm": float(output) if isinstance(output, (int, float)) else 0.0,
        }
    wb.close()
    return out


# -----------------------------------------------------------------------------
# Load energy-product imports per industry from the product workbook's
# 'Imports use pxi' table, plus the aggregate import intensity (imports as
# share of total intermediate input spending) per industry from the IOT
# sheet's totals rows.
#
# The first three values become the A_imp[energy, j] entries that we add
# to the corrected energy-intensity vector at each domestic node (see
# build()). The aggregate import intensity is a flag for which industries
# have the largest unmeasured upstream supply chain — i.e. where embodied
# energy in imported NON-energy intermediates (steel, chemicals, etc.) is
# most likely to be material and is NOT picked up by the domestic Leontief.
# -----------------------------------------------------------------------------
ENERGY_IMPORT_CPAS = {
    "CPA_C19":     "imp_oil_gbpm",
    "CPA_D351":    "imp_elec_gbpm",
    "CPA_D352_3":  "imp_gas_gbpm",
}

def load_imports_and_intermediates():
    out = {}

    # 1. Per-industry energy-product imports from product workbook
    wb_p = openpyxl.load_workbook(XLSX_PRODUCT, data_only=True, read_only=True)
    ws = wb_p["Imports use pxi"]
    col_codes = {}
    for c in range(3, ws.max_column + 1):
        code = ws.cell(4, c).value
        if code is None:
            break
        col_codes[c] = str(code).strip()
    energy_rows = {}
    for r in range(7, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v in ENERGY_IMPORT_CPAS:
            energy_rows[v] = r
    missing = set(ENERGY_IMPORT_CPAS) - set(energy_rows)
    assert not missing, f"missing energy CPAs in Imports use pxi: {missing}"

    for c, code in col_codes.items():
        rec = {"imp_oil_gbpm": 0.0, "imp_gas_gbpm": 0.0, "imp_elec_gbpm": 0.0}
        for cpa, key in ENERGY_IMPORT_CPAS.items():
            v = ws.cell(energy_rows[cpa], c).value
            rec[key] = float(v) if isinstance(v, (int, float)) else 0.0
        out[code] = rec
    wb_p.close()

    # 2. Aggregate domestic intermediate use + total imports from the
    #    industry workbook (rows 111 and 112 of the IOT sheet).
    wb_i = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    ws = wb_i["IOT"]
    for c in range(3, ws.max_column + 1):
        code = ws.cell(4, c).value
        if code is None:
            break
        code = str(code).strip()
        rec = out.setdefault(code, {"imp_oil_gbpm": 0.0, "imp_gas_gbpm": 0.0, "imp_elec_gbpm": 0.0})
        dom = ws.cell(111, c).value
        imp = ws.cell(112, c).value
        rec["dom_intermediate_gbpm"] = float(dom) if isinstance(dom, (int, float)) else 0.0
        rec["imp_total_gbpm"]        = float(imp) if isinstance(imp, (int, float)) else 0.0
        denom = rec["dom_intermediate_gbpm"] + rec["imp_total_gbpm"]
        rec["import_intensity"] = (rec["imp_total_gbpm"] / denom) if denom else 0.0
    wb_i.close()
    return out


# -----------------------------------------------------------------------------
# Load A matrix from ONS xlsx
# -----------------------------------------------------------------------------
def load_A():
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    ws = wb["A"]
    # Row 5 (1-indexed) has column codes starting at col 3; col A code labels
    # start at row 8. We read until column/row label goes blank.
    col_codes, col_names = [], []
    for c in range(3, ws.max_column + 1):
        code = ws.cell(5, c).value
        name = ws.cell(6, c).value
        if code is None:
            break
        col_codes.append(str(code).strip())
        col_names.append(str(name).strip() if name else "")
    row_codes, row_names = [], []
    for r in range(8, ws.max_row + 1):
        code = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        if code is None:
            break
        row_codes.append(str(code).strip())
        row_names.append(str(name).strip() if name else "")
    # IxI tables are square
    n = min(len(row_codes), len(col_codes))
    A = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            v = ws.cell(8 + i, 3 + j).value
            A[i, j] = float(v) if isinstance(v, (int, float)) else 0.0
    codes = row_codes[:n]
    names = row_names[:n]
    # Sanity: column codes should match row codes
    assert codes == col_codes[:n], "IxI rows and columns should align"
    return A, codes, names


# -----------------------------------------------------------------------------
# Sector-specific realistic pass-through coefficients
# Calibrated qualitatively to trade exposure + market power per the literature.
#   >=1.00 : concentrated, domestic-oriented, can expand margin (LMM 2023: ~115%
#            in least-competitive sectors)
#   0.90-1.00 : strong pricing power / differentiated output (food, pharma,
#               industrial gases)
#   0.60-0.80 : competitive but somewhat protected (cement, rubber & plastics)
#   0.30-0.50 : highly trade-exposed commodity manufacturing (steel, aluminium,
#               basic chemicals, paper, glass, ceramics) — Ganapati et al.
#               (2020) avg ~70%, tradeable commodity tail below that
# -----------------------------------------------------------------------------
# -----------------------------------------------------------------------------
# Shorthand industry names. ONS publishes long, formal SIC titles which are
# unwieldy in tables (e.g. "Manufacture of glass, refractory, clay, porcelain,
# ceramic, stone products - 23.1-4/7-9"). The shorthand below is what gets
# emitted to irf.json and rendered in the table; the original ONS title is
# preserved alongside as `name_full` for reference.
# -----------------------------------------------------------------------------
NAME_OVERRIDES = {
    "A01":              "Agriculture",
    "A02":              "Forestry",
    "A03":              "Fishing & aquaculture",
    "B05":              "Coal mining",
    "B06 & B07":        "Oil/gas extraction & metal ores",
    "B08":              "Other mining",
    "B09":              "Mining support services",
    "C101":             "Meat processing",
    "C102_3":           "Fish & fruit/veg processing",
    "C104":             "Edible oils & fats",
    "C105":             "Dairy",
    "C106":             "Grain milling & starches",
    "C107":             "Bakery",
    "C108":             "Other food mfg",
    "C109":             "Animal feeds",
    "C1101T1106 & C12": "Alcohol & tobacco",
    "C1107":            "Soft drinks",
    "C13":              "Textiles",
    "C14":              "Apparel",
    "C15":              "Leather",
    "C16":              "Wood & wood products",
    "C17":              "Paper & paper products",
    "C18":              "Printing",
    "C19":              "Refined petroleum",
    "C203":             "Paints & coatings",
    "C204":             "Soaps & detergents",
    "C205":             "Other chemicals",
    "C20A":             "Industrial gases / fertilisers",
    "C20B":             "Petrochemicals",
    "C20C":             "Dyestuffs & agrochemicals",
    "C21":              "Pharmaceuticals",
    "C22":              "Rubber & plastics",
    "C235_6":           "Cement, lime & concrete",
    "C23OTHER":         "Glass & ceramics",
    "C241T243":         "Iron & steel",
    "C244_5":           "Aluminium & other basic metals",
    "C25":              "Fabricated metal",
    "C26":              "Electronics & optical",
    "C27":              "Electrical equipment",
    "C28":              "Machinery",
    "C29":              "Motor vehicles",
    "C301":             "Shipbuilding",
    "C303":             "Aerospace mfg",
    "C30OTHER":         "Other transport equipment",
    "C31":              "Furniture",
    "C32":              "Other manufacturing",
    "C3315":            "Ship repair",
    "C3316":            "Aircraft repair",
    "C33OTHER":         "Other repair & install",
    "D351":             "Electricity",
    "D352_3":           "Gas & steam",
    "E36":              "Water supply",
    "E37":              "Sewerage",
    "E38":              "Waste management",
    "E39":              "Remediation",
    "F41, F42  & F43":  "Construction",
    "G45":              "Vehicle wholesale & retail",
    "G46":              "Wholesale (non-motor)",
    "G47":              "Retail (non-motor)",
    "H491_2":           "Rail transport",
    "H493T495":         "Land transport (road + pipelines)",
    "H50":              "Water transport",
    "H51":              "Air transport",
    "H52":              "Warehousing",
    "H53":              "Postal & courier",
    "I55":              "Accommodation",
    "I56":              "Restaurants & food service",
    "J58":              "Publishing",
    "J59 & J60":        "TV, film & sound",
    "J61":              "Telecoms",
    "J62":              "IT services",
    "J63":              "Information services",
    "K64":              "Banking",
    "K65.1-2 & K65.3":  "Insurance & pensions",
    "K66":              "Financial auxiliary",
    "L683":             "Real estate (fee-based)",
    "L68A":             "Owner-occupier housing",
    "L68BXL683":        "Real estate (own/lease)",
    "M691":             "Legal",
    "M692":             "Accounting & audit",
    "M70":              "Management consultancy",
    "M71":              "Architecture & engineering",
    "M72":              "R&D",
    "M73":              "Advertising",
    "M74":              "Other professional services",
    "M75":              "Veterinary",
    "N77":              "Rental & leasing",
    "N78":              "Employment services",
    "N79":              "Travel agencies",
    "N80":              "Security services",
    "N81":              "Building services",
    "N82":              "Office support",
    "O84":              "Public admin & defence",
    "P85":              "Education",
    "Q86":              "Human health",
    "Q87 & Q88":        "Social care",
    "R90":              "Arts & entertainment",
    "R91":              "Libraries & museums",
    "R92":              "Gambling",
    "R93":              "Sports & recreation",
    "S94":              "Membership organisations",
    "S95":              "Personal-goods repair",
    "S96":              "Personal services",
    "T97":              "Households as employers",
}
def short_name(code: str, fallback: str) -> str:
    return NAME_OVERRIDES.get(code, fallback.strip())


# -----------------------------------------------------------------------------
# DESNZ Energy & Trade Intensive Industries (ETII) eligibility, mapped from
# the published 4-digit SIC list (DESNZ Jan 2023 ETII methodology PDF, Annex A)
# onto the IOAT 2023 industry buckets used in this dataset. Three tiers:
#
#   "etii"    — the IOAT bucket is wholly or predominantly on the ETII list
#   "partial" — some sub-codes inside the bucket are ETII, others are not
#   None      — none of the bucket is ETII
#
# Where the IOAT bucket aggregates multiple 4-digit codes the precision is
# inevitably partial; the table tooltip flags this. A 4-digit-precise
# version would require working at the SIC class level using ABS/ECUK data.
# -----------------------------------------------------------------------------
ETII_STATUS = {
    # Full ETII buckets
    "C17":      "etii",   # 17.11/12/21-29  pulp & paper
    "C20A":     "etii",   # 20.11/13/15     industrial gases / inorganics / fertilisers
    "C20B":     "etii",   # 20.14/16/17/60  organic chemicals / plastics primary / synthetic rubber / man-made fibres
    "C20C":     "etii",   # 20.12/20        dyestuffs / agrochemicals
    "C235_6":   "etii",   # 23.51/52 + 61-69  cement, lime, concrete
    "C23OTHER": "etii",   # 23.11-20/31-49/99  glass, refractories, ceramics, abrasives
    "C241T243": "etii",   # 24.10/20/31-34  iron & steel
    "C244_5":   "etii",   # 24.41-46/51-54  aluminium + other basic metals + casting
    # Partial ETII buckets (some sub-codes on the list, others not)
    "B06 & B07": "partial",  # 07.x metal ores ETII; 06.x oil/gas extraction not
    "B08":       "partial",  # 08.91/93/99 ETII; 08.11/12/etc not
    "C104":      "partial",  # 10.41 oils & fats ETII; 10.42 not
    "C106":      "partial",  # 10.62 starches ETII; 10.61 grain milling not
    "C108":      "partial",  # 10.81 sugar ETII; rest of bucket not
    "C1101T1106 & C12": "partial",  # 11.06 malt ETII; rest of bucket not
    "C13":       "partial",  # 13.10/20/30/94/95/96 ETII; 13.91/92/93/99 not
    "C16":       "partial",  # 16.10 sawmilling, 16.21 veneer ETII; 16.22-29 not
    "C205":      "partial",  # 20.59 partly ETII; 20.51-53 not
    "C22":       "partial",  # 22.11/19/21/22 ETII; 22.23/29 not
}


# Sector-specific pass-through coefficients ρ_j. Calibrated against:
#   - Ganapati, Shapiro & Walker (AEJ Applied 2020): US manufacturing average
#     ~70%, range from <70% in commodity / trade-exposed segments to >100%
#     in concentrated industries.
#   - Lafrogne-Joussier, Martin & Méjean (CEPII 2023): French firm-level,
#     ~100% on energy costs in the 2020–22 episode, ~115% in least-competitive
#     sectors. Channel ranking: cut energy → raise prices → switch imports →
#     shift plants → invest in efficiency.
#   - Bank of England / Bank Underground (2023): UK PPI/SPPI sector
#     regressions 1997+. Long-run input-cost pass-through ~0.8 for
#     manufacturing, ~0.4 for services. Manufacturing reaches 80% of the
#     long-run pass-through in ~8 quarters; services in ~15. Asymmetry of
#     ~2 quarters longer when costs are falling.
#   - Cambridge EPRG NTS1935 / CE Delft / Öko-Institut (2019–20): EU ETS
#     ex-post pass-through bands by sector — refining 80–100%, iron &
#     steel 55–85%, cement 20–40%, chemicals/fertilisers high but
#     uncertain. Ex-post additional EU industrial profits from carbon
#     pass-through estimated at €26–46bn over 2008–2019.
#
# UK-specific overlay: trade exposure (steel exports ~45%, imports >60%
# of demand; ETII list) and market-power judgements pull commodity-EII
# values to the bottom of the international bands; transport-protected
# domestic sectors (cement, food) get higher ρs reflecting LMM and
# UK-observed pass-through during 2022.
RHO_OVERRIDES = {
    # Heavily trade-exposed EII commodity manufacturing
    "C241T243": 0.50,   # Basic iron & steel — centre of EU ETS 0.55–0.85 band; UK at low end given Tata closures
    "C244_5":   0.50,   # Other basic metals (incl. aluminium) — same reasoning
    "C20A":     0.40,   # Industrial gases, inorganics, fertilisers
    "C20B":     0.65,   # Petrochemicals — EU ETS / chemicals literature suggests mid-band
    "C20C":     0.55,   # Dyestuffs, agrochemicals
    "C17":      0.55,   # Paper & paper products
    "C235_6":   0.80,   # Cement, lime, plaster (transport-protected domestically; consistent with BoE 0.8 mfg)
    "C23OTHER": 0.50,   # Glass, ceramics, refractories
    "C13":      0.55,   # Textiles
    "C14":      0.55,   # Wearing apparel
    "C15":      0.50,   # Leather
    "C16":      0.60,   # Wood & wood products
    "C22":      0.70,   # Rubber & plastics
    "C19":      0.90,   # Refined petroleum (EU ETS literature: refining 80–100%; shocked directly)
    # Concentrated / differentiated / strong pricing power
    "C101":     1.00, "C102_3": 1.00, "C104": 1.00, "C105": 1.00,
    "C106":     1.00, "C107": 1.00, "C108": 1.00, "C109": 1.00,
    "C1101T1106 & C12": 1.00, "C1107": 1.00,   # Food & beverages — LMM + ONS 2025
    "C21":      0.95,  # Pharmaceuticals
    "D351":     1.00,  # Electricity (regulated pass-through)
    "D352_3":   1.00,
    # Utilities / non-tradeable
    "E36": 1.00, "E37": 1.00, "E38": 0.95, "E39": 0.95,
    "F41, F42  & F43": 0.90,  # Construction
    # Services default 0.90: anchored on LMM's energy-specific finding
    # rather than BoE's 0.4 (which is long-run for general input costs;
    # energy is a faster-passing supply-side shock).
}
def rho_for(code: str) -> float:
    return RHO_OVERRIDES.get(code, 0.90)


# -----------------------------------------------------------------------------
# Build IRF
# -----------------------------------------------------------------------------
def build():
    A, codes, names = load_A()
    n = len(codes)
    idx = {c: i for i, c in enumerate(codes)}
    gva_out = load_gva_and_output()
    imports = load_imports_and_intermediates()

    # Corrected energy-intensity vectors per £ of each industry's output:
    # e_<fuel>[i] = A_dom[fuel, i] + A_imp[fuel, i]. Built once here so the
    # direct, total, indirect and via_transport calculations in the per-row
    # loop below all share the same numerator at each node. Imported
    # non-energy intermediates are NOT added — they would require a
    # multi-region IO model to propagate properly.
    e_gas  = np.zeros(n)
    e_elec = np.zeros(n)
    e_oil  = np.zeros(n)
    for i, code in enumerate(codes):
        meta_imp = imports.get(code, {})
        out_gbpm = gva_out.get(code, {}).get("output_gbpm", 0.0)
        a_imp_gas  = (meta_imp.get("imp_gas_gbpm",  0.0) / out_gbpm) if out_gbpm else 0.0
        a_imp_elec = (meta_imp.get("imp_elec_gbpm", 0.0) / out_gbpm) if out_gbpm else 0.0
        a_imp_oil  = (meta_imp.get("imp_oil_gbpm",  0.0) / out_gbpm) if out_gbpm else 0.0
        e_gas[i]  = float(A[idx["D352_3"], i]) + a_imp_gas
        e_elec[i] = float(A[idx["D351"],   i]) + a_imp_elec
        e_oil[i]  = float(A[idx["C19"],    i]) + a_imp_oil

    e_idx = [idx[c] for c in ENERGY_CODES if c in idx]
    r_idx = [i for i in range(n) if i not in e_idx]
    missing = [c for c in ENERGY_CODES if c not in idx]
    assert not missing, f"missing energy codes in IxI: {missing}"

    Arr = A[np.ix_(r_idx, r_idx)]   # R x R — non-energy block, domestic-A only
    # Imports-corrected energy → non-energy block: each col j shows TOTAL
    # (UK + imp) energy inputs to non-energy industry j per £ of j's output.
    # Built from the e_<fuel> vectors above so the same correction that
    # feeds the exposures table also feeds the IRF — imported jet kerosene
    # at H51 propagates to industries buying air transport via Aer; the
    # downstream Arr propagation stays domestic-only because we don't have
    # foreign A for non-energy intermediates.
    e_per_code = {"C19": e_oil, "D351": e_elec, "D352_3": e_gas}
    e_stacked  = np.vstack([e_per_code[c] for c in ENERGY_CODES])  # E x n
    Aer        = e_stacked[:, r_idx]                                # E x R

    # Long-run (full pass-through) closed-form:
    # dp_R = (I - Arr')^{-1} Aer' dp_E
    I = np.eye(len(r_idx))
    L_rr = np.linalg.inv(I - Arr.T)  # (R x R)

    def run(shock_dict, apply_rho=False):
        dp_E = np.array([shock_dict[ENERGY_CODES[k]] for k in range(len(e_idx))])
        # direct impulse to R: Aer' . dp_E (R-vector)
        direct = Aer.T @ dp_E
        # dynamic path
        quarters = np.zeros((HORIZON_QUARTERS + 1, len(r_idx)))
        quarters[0] = 0.0
        running = np.zeros(len(r_idx))
        for q in range(1, HORIZON_QUARTERS + 1):
            # p_R^{(q)} = direct + Arr' p_R^{(q-1)}
            new_running = direct + Arr.T @ running
            if apply_rho:
                rho_vec = np.array([rho_for(codes[r_idx[k]]) for k in range(len(r_idx))])
                # Apply rho at each round as a dampener on the running response
                # Equivalent to: effective pass-through = rho * full
                new_running = rho_vec * new_running
            quarters[q] = new_running
            running = new_running
        # Closed-form long-run (full path only — "realistic" stays at its q=12 value)
        lr = L_rr @ direct
        return direct, quarters, lr

    # Full-economy Leontief inverse L = (I - A)^-1 (includes the energy
    # rows, unlike L_rr above). Used to compute total (direct + indirect)
    # energy intensity per £1 of an industry's output, and to decompose
    # indirect oil exposure by transport channel.
    L_full = np.linalg.inv(np.eye(n) - A)

    transport_idx = [idx[c] for c in TRANSPORT_CODES if c in idx]

    out_rows = []
    # Per-industry energy cost shares using corrected (UK + imp) energy-
    # intensity vectors at each domestic node, propagated through the
    # domestic Leontief L_full = (I − A_dom)^-1. Direct = e_*[j] is the
    # total energy bought by j (UK + imports). Total = e_* · L_full[:, j]
    # is energy embedded along all UK supply chains, with each node's
    # energy use measured correctly. Indirect = total − direct.
    # NOTE: this still excludes embodied energy in imported NON-energy
    # intermediates (foreign supply chains we don't have data for). Per-
    # industry import_intensity flag below surfaces where that gap matters.
    for i in range(n):
        if i in e_idx:
            continue
        code = codes[i]
        name = names[i]

        # Direct (UK + imported) energy bought by j per £ output
        gas   = float(e_gas[i])
        elec  = float(e_elec[i])
        oil   = float(e_oil[i])
        direct_share = gas + elec + oil

        # Total energy embedded via UK supply chains, with corrected
        # energy intensity at each node: total[fuel, j] = e_fuel · L[:, j]
        total_gas  = float(e_gas  @ L_full[:, i])
        total_elec = float(e_elec @ L_full[:, i])
        total_oil  = float(e_oil  @ L_full[:, i])

        # Indirect = total − direct (the indirect part propagates through
        # UK domestic chains; each upstream node's e_fuel includes its own
        # imports of energy, so airline-bought kerosene shows up when any
        # downstream industry buys air-transport services).
        indirect_gas  = total_gas  - gas
        indirect_elec = total_elec - elec
        indirect_oil  = total_oil  - oil

        # Indirect oil arriving with any transport industry as last hop.
        # Uses e_oil at each transport node so imported jet kerosene /
        # marine bunker / road diesel embedded in transport services bought
        # by j is captured.
        # Diagonal fix: when j IS a transport industry, the T=j term of
        # the sum contains e_oil[j]·L[j, j]. The leading "1" of L[j, j] is
        # j's own direct purchase — already in `oil`. Subtract it so
        # direct + via_transport + non_transport remains a clean partition.
        oil_via_transport = sum(
            float(e_oil[ti]) * float(L_full[ti, i]) for ti in transport_idx
        )
        if i in transport_idx:
            oil_via_transport -= float(e_oil[i])
        # Residual: indirect oil through non-transport intermediates
        # (chemicals, plastics, basic metals, agriculture etc.).
        oil_indirect_non_transport = indirect_oil - oil_via_transport

        # GVA conversion factor: a share-of-output × (output / GVA) = share-of-GVA.
        # Applied identically to direct and total because both are denominated
        # by gross output at basic prices in the A and L matrices.
        meta   = gva_out.get(code, {})
        gva    = meta.get("gva_gbpm",    0.0)
        output = meta.get("output_gbpm", 0.0)
        gva_ratio = (output / gva) if gva else None  # output / GVA — multiplier

        # Import flag: share of j's total intermediate input spending that is
        # imported. Industries with high values have a meaningful upstream
        # supply-chain undercount (foreign embodied energy in imported non-
        # energy intermediates is NOT in the domestic Leontief).
        imp_meta = imports.get(code, {})
        import_intensity = imp_meta.get("import_intensity", 0.0)

        def to_gva(x):
            return (x * gva_ratio) if (gva_ratio is not None and x is not None) else None

        # ETII-style direct gas+elec / GVA cost share — DESNZ uses the 80th
        # percentile of this metric as its eligibility cutoff.
        gas_elec_direct_gva   = to_gva(gas + elec)
        # Total (direct + indirect) gas+elec+oil / GVA — the comparable
        # all-fuels share-of-GVA used in most academic energy-intensity work.
        total_share_gva       = to_gva(direct_share + indirect_gas + indirect_elec + indirect_oil)
        # Direct all-fuels / GVA — useful side-by-side with ETII (electricity+gas
        # only) to see how much oil adds to the direct share.
        direct_share_gva      = to_gva(direct_share)

        out_rows.append({
            "code": code,
            "name":      short_name(code, name),
            "name_full": name.strip(),
            # Direct (column of A) — share of gross output
            "direct_gas":   gas,
            "direct_elec":  elec,
            "direct_oil":   oil,
            "direct_share": direct_share,
            # Total embodied (column of L) — share of gross output
            "total_gas":   total_gas,
            "total_elec":  total_elec,
            "total_oil":   total_oil,
            "total_share": total_gas + total_elec + total_oil,
            # Indirect = total − direct
            "indirect_gas":  indirect_gas,
            "indirect_elec": indirect_elec,
            "indirect_oil":  indirect_oil,
            # Channel decomposition of indirect oil
            "oil_via_transport":          oil_via_transport,
            "oil_indirect_non_transport": oil_indirect_non_transport,
            # GVA + share-of-GVA versions for literature comparison
            "gva_gbpm":             gva,
            "output_gbpm":          output,
            "gas_elec_direct_gva":  gas_elec_direct_gva,
            "direct_share_gva":     direct_share_gva,
            "total_share_gva":      total_share_gva,
            # Import intensity of intermediate inputs — flag for the
            # residual MRIO gap (foreign embodied energy in imported non-
            # energy intermediates, not captured by domestic Leontief).
            "import_intensity":     import_intensity,
            # DESNZ ETII flag (etii / partial / null) — see ETII_STATUS above.
            "etii":                 ETII_STATUS.get(code),
            "imp_total_gbpm":       imp_meta.get("imp_total_gbpm", 0.0),
            "imp_oil_gbpm":         imp_meta.get("imp_oil_gbpm",   0.0),
            "imp_gas_gbpm":         imp_meta.get("imp_gas_gbpm",   0.0),
            "imp_elec_gbpm":        imp_meta.get("imp_elec_gbpm",  0.0),
        })

    # Run scenarios
    scen_defs = [
        ("unit_full",       SHOCK_UNIT,   False),
        ("unit_realistic",  SHOCK_UNIT,   True),
        ("stress_full",     SHOCK_STRESS, False),
        ("stress_realistic",SHOCK_STRESS, True),
    ]
    scenarios = {}
    for name_, shock, rho in scen_defs:
        direct, path, lr = run(shock, apply_rho=rho)
        scenarios[name_] = {"direct": direct, "path": path, "lr": lr}

    # Attach IRF per industry
    r_code_order = [codes[i] for i in r_idx]
    code_to_ridx = {c: k for k, c in enumerate(r_code_order)}
    for row in out_rows:
        k = code_to_ridx[row["code"]]
        row["rho"] = rho_for(row["code"])
        for sname in scenarios:
            sc = scenarios[sname]
            row[f"{sname}_irf"] = [float(x) for x in sc["path"][:, k]]
            row[f"{sname}_lr"]  = float(sc["lr"][k])
            row[f"{sname}_direct"] = float(sc["direct"][k])

    # Top-level metadata
    result = {
        "source": "ONS UK Input-Output Analytical Tables (Industry by Industry), reference year 2023; "
                  "Blue Book 2025 vintage; file data/iot2023industry.xlsx.",
        "energy_codes": ENERGY_CODES,
        "transport_codes": TRANSPORT_CODES,
        "n_industries": n,
        "horizon_quarters": HORIZON_QUARTERS,
        "shock_unit":   SHOCK_UNIT,
        "shock_stress": SHOCK_STRESS,
        "methodology": (
            "Leontief price-side IRF: treat energy industries (C19, D351, D352_3) "
            "as exogenous and compute dp_R = (I - A_RR')^-1 A_ER' dp_E for the "
            "full pass-through long-run limit. The energy → non-energy block "
            "A_ER uses the IMPORTS-CORRECTED energy-intensity vector at each "
            "domestic node, e_fuel[j] = A_dom[fuel,j] + A_imp[fuel,j], so "
            "imported jet kerosene / marine bunker / diesel feed the price "
            "response of any industry that buys them directly or buys "
            "transport services from an industry that does. The non-energy → "
            "non-energy block A_RR stays domestic-A only. Dynamic path built "
            "by Neumann series with one round = one quarter. Realistic "
            "scenarios apply sector-specific rho_j pass-through coefficients "
            "(calibrated from Ganapati/Shapiro/Walker 2020 and Lafrogne-"
            "Joussier/Martin/Mejean 2023). Foreign embodied energy in "
            "imported non-energy intermediates is NOT captured — see the "
            "import_intensity flag for which industries that gap matters most."
        ),
        "industries": out_rows,
    }
    OUT.write_text(json.dumps(result, indent=1))
    print(f"wrote {OUT} — {len(out_rows)} industries")


if __name__ == "__main__":
    build()
